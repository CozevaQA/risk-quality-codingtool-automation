import os
import random
from statistics import mean
import sys
import math
import time
import traceback
import csv
from os import listdir
from os.path import isfile, join
import re
from difflib import SequenceMatcher
import json
import datetime
import base64

from pandas.core.config_init import pc_max_info_rows_doc
from selenium.webdriver.common.action_chains import ActionChains
from selenium import *
from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.support import expected_conditions as EC
from selenium.common.exceptions import *
from selenium.webdriver.support.wait import WebDriverWait
from openpyxl import Workbook
from selenium.webdriver import Keys
from selenium.webdriver.support.ui import Select
from CodeCheckUI import launch_measure_selector
from openpyxl.styles import Font, Alignment, PatternFill
from openpyxl.utils import get_column_letter
from urllib.parse import urlparse, parse_qs
import random



result = launch_measure_selector()



def ajax_preloader_wait(driver):
    time.sleep(1)
    WebDriverWait(driver, 300).until(EC.invisibility_of_element((By.CLASS_NAME, "ajax_preloader")))
    if len(driver.find_elements(By.CLASS_NAME, "ajax_preloader")) != 0:
        WebDriverWait(driver, 300).until(EC.invisibility_of_element((By.CLASS_NAME, "ajax_preloader")))
    WebDriverWait(driver, 300).until(EC.invisibility_of_element((By.CLASS_NAME, "drupal_message_text")))
    time.sleep(1)

def measure_search_registry(driver, measure):
    driver.find_element(By.XPATH, "//a[@data-target='qt-reg-nav-filters']").click()
    time.sleep(0.5)
    driver.find_element(By.XPATH, "//*[@id='qt-reg-nav-filters']/li[1]/label").click()
    time.sleep(0.5)
    driver.find_element(By.XPATH, "//*[@id='qt-search-met']").send_keys(measure)
    apply_btn = driver.find_element(By.XPATH, "//button[@id='qt-apply-search']")
    driver.execute_script("arguments[0].scrollIntoView();", apply_btn)
    apply_btn.click()
    time.sleep(0.5)

def measure_patient_finder(driver, customer, lob, measure_name, measure_abb,domain_name, ws, wb):
    measure_link = ""
    name_check = "NA"
    domain_check = "NA"
    pencil_check = "NA"
    pencil_icon_options = []
    pencil_icon_details = "NA"
    perf_check = "NA"
    network_check = "NA"
    dashboard_map_check = "NA"
    dashboard_pencil_check = "NA"
    patient_link = "NA"
    comments = "NA"
    mspl_dashboard_check = ""
    mspl_map_flag = 0
    mspl_supp_flag = 0
    dash_map_flag = 0
    dash_supp_flag = 0
    compliancy_details = "NA"
    mspl_link = ""
    try:
        #measure_link = driver.find_element(By.XPATH, "//*[contains(text(),'"+measure_abb+"')]//..//..//..//..").get_attribute('href')
        try:
            measure_link = driver.find_element(By.XPATH,
                                               "//*[contains(text(),'" + measure_name + "')]//..//..//..//..").get_attribute(
                'href')
            print("Test....6")
        except Exception as e:
            split_text = measure_name.split("'", 1)
            modified_text = split_text[1]
            print("Test....7")
            measure_link = driver.find_element(By.XPATH,
                                               "//span[contains(text(),'" + modified_text + "')]//ancestor::a").get_attribute(
                'href')
        driver.get(measure_link)
        ajax_preloader_wait(driver)
        WebDriverWait(driver, 30).until(EC.visibility_of_element_located((By.XPATH, "//*[@class='dataTables_info']")))
        measure_fullname = driver.find_element(By.XPATH,"//*[contains(@class,'metric_specific_patient_list_title')]").text
        measure_domain_ui = measure_fullname.split("|")[0]
        measure_name_ui = measure_fullname.split("|")[1]
        print("Sheet: "+domain_name)
        print("UI: "+measure_domain_ui)
        if measure_name in measure_name_ui:
            name_check = "Measure name successfully validated"
        else:
            name_check = "Measure name discrepancy found"
        if domain_name in measure_domain_ui:
            domain_check = "Domain name successfully validated"
        else:
            domain_check = "Domain name discrepancy found"
        try:
            elems = driver.find_element(By.XPATH, "//*[@id='metric-support-prov-ls']/tbody/tr[1]//td[5]").find_elements(By.TAG_NAME, 'a')
            mspl_link = elems[0].get_attribute('href')
            driver.get(mspl_link)
            print("Opening MSPL...")
            ajax_preloader_wait(driver)
            WebDriverWait(driver, 30).until(EC.visibility_of_element_located((By.XPATH, "//*[@class='dataTables_info']")))
            driver.find_element(By.XPATH, "//*[@class='dt_tag_wrapper']//*[contains(@class,'dt_tag_close')]").click()
            ajax_preloader_wait(driver)
            WebDriverWait(driver, 30).until(EC.visibility_of_element_located((By.XPATH, "//*[@class='dataTables_info']")))
            rows = driver.find_element(By.XPATH, "//*[@id='quality_registry_list']").find_element(By.TAG_NAME, 'tbody').find_elements(By.TAG_NAME, 'tr')
            patient_link = rows[0].find_element(By.XPATH, "//*[contains(@class,'pat_name')]").get_attribute('href')
            tab_param = driver.find_element(By.XPATH, "//*[@id='quality_registry']//ul[@class='tabs']").find_elements(By.TAG_NAME, 'li')
            perf_check = tab_param[1].text
            network_check = tab_param[2].text
            try:
                print("Checking pencil icon availability...")
                rows[0].find_element(By.XPATH, "//*[contains(@class,'addSuppData-trigger')]").click()
                time.sleep(3)
                pencil_icon_details = rows[0].find_element(By.XPATH, "//*[contains(@class,'addSuppData-trigger')]//..//ul").find_elements(By.TAG_NAME, 'li')
                pencil_check = "Pencil icon available"
                for detail in pencil_icon_details:
                    string = detail.text
                    pencil_icon_options.append(string)
                pencil_icon_details = ", ".join(pencil_icon_options)
            except Exception as e:
                pencil_check = "Pencil icon not available"
                pencil_icon_details = "Not applicable"
        except NoSuchElementException as e:
            print(e)
            print("No Providers Found, skipping check")
            comments = "No Providers Found for the measure"
            pencil_icon_details = "NA"
            patient_link = "NA"
    except Exception as e:
        print("Error occured!!!")
    if patient_link == "NA":
        print("Skipping to next check...")
        dashboard_check = "No dashboard to check"
    else:
        driver.get(patient_link)
        ajax_preloader_wait(driver)
        measure_name_dashboard = measure_name.replace("'","")
        try:
            driver.find_element(By.XPATH, "//*[contains(text(),'" + measure_name_dashboard + "')]//..//..//..//..//*[contains(@class,'pre_process_hcc')]")
            dashboard_pencil_check = "Pencil icon available"
            dash_supp_flag = 1
        except NoSuchElementException as e:
            dashboard_pencil_check = "Pencil icon not available"
        try:
            driver.find_element(By.XPATH, "//*[contains(text(),'" + measure_name_dashboard + "')]//..//..//..//..//*[@class='toggle_patient_metric_pending']")
            dashboard_map_check = "MAP checkbox available"
            dash_map_flag = 1
        except NoSuchElementException as e:
            dashboard_map_check = "MAP checkbox available"
        try:
            driver.find_element(By.XPATH, "//*[contains(text(),'"+measure_name+"')]//..//..//..//..//*[contains(@class,'non_compliant')]")
            compliancy_details = "Compliant"
        except NoSuchElementException as e:
            compliancy_details = "Non-compliant"
    if "Add Supplemental Data" in pencil_icon_details:
        mspl_supp_flag = 1
    else:
        comments = comments + "| No supplemental data, skipping code check"
    if "Mark as Pending" in pencil_icon_details:
        mspl_map_flag = 1
    if mspl_map_flag == dash_map_flag and mspl_supp_flag == dash_supp_flag:
        mspl_dashboard_check = "MATCH"
    else:
        mspl_dashboard_check = "NOT MATCH"
    ws.append([customer, lob, domain_check, name_check, perf_check, network_check, comments])
    ws = wb["MSPL vs Dashboard Check"]
    ws.append([customer, lob,compliancy_details, pencil_check, pencil_icon_details, dashboard_pencil_check, dashboard_map_check,mspl_dashboard_check, mspl_link])
    return mspl_supp_flag


def normalize(text):
    text = re.sub(r"\d+[-–]\d+", "", text)
    text = " ".join(text.lower().split())
    return text

def supplemental_data_launcher(driver):
    coding_tool_btn = driver.find_element(By.XPATH, "//*[contains(text(),'Add Supplemental Data')]")
    driver.execute_script("arguments[0].scrollIntoView({block:'center', inline:'nearest'});", coding_tool_btn)
    coding_tool_btn.click()
    ajax_preloader_wait(driver)
    WebDriverWait(driver, 30).until(EC.visibility_of_element_located((By.XPATH, "//*[contains(@class,'task_def')]")))

def code_search_coding_tool(driver, customer, lob, measure_name, ws):
    criteria = ""
    result = ""
    populated_result = ""
    cz_id = driver.find_element(By.XPATH, "//*[contains(@class,'patient_cozeva_id')]").text
    time.sleep(2)
    metric_rows = driver.find_elements(By.XPATH, "//*[@class='metric']")
    i=1
    try:
        print("Finding matching metric...")
        for i in range(1, len(metric_rows),1):
            measure_name_ui = metric_rows[i].find_elements(By.XPATH, "//*[contains(@class,'met-name')]")[i].text
            if measure_name == measure_name_ui:
                print("Matched metric: "+measure_name_ui)
                driver.execute_script("arguments[0].scrollIntoView({block:'center', inline:'nearest'});",metric_rows[i])
                driver.find_element(By.XPATH, "//*[@id='right_panel_nav_tabs']/li[1]/div").click()
                time.sleep(1)
                metric_rows[i].find_elements(By.XPATH, "//*[contains(@class,'add_more')]")[i].click()
                time.sleep(2)
                break
    except Exception as e:
        print(e)
        print("Error Occured!")
    # driver.find_element(By.XPATH, "//*[contains(@class,'choice_tr')]//*[@class='choice_option']//*[@class='select-wrapper']").click()
    choices = driver.find_element(By.XPATH, "//*[contains(@class,'choice_tr')]//*[@class='choice_option']//ul").find_elements(By.TAG_NAME, 'li')
    print("Executing code search...")
    for i in range(0, len(choices),1):
        driver.find_element(By.XPATH, "//*[contains(@class,'choice_tr')]//*[@class='choice_option']//*[@class='select-wrapper']").click()
        time.sleep(2)
        choices = driver.find_element(By.XPATH,
                                      "//*[contains(@class,'choice_tr')]//*[@class='choice_option']//ul").find_elements(
            By.TAG_NAME, 'li')
        criteria = choices[i].text
        print("Checking Criteria "+str(i+1)+": "+criteria)
        time.sleep(1)
        choices[i].click()
        time.sleep(3)
        result_list = []
        try:
            WebDriverWait(driver, 30).until(EC.visibility_of_element_located((By.XPATH, "//*[@class='advanced_search_btn']")))
            driver.find_element(By.XPATH, "//*[@class='advanced_search_btn']").click()
        except NoSuchElementException as e:
            ws.append([customer, lob,cz_id, measure_name, criteria, "No codes for criteria","NA","NA"])
            continue
        time.sleep(1)
        populated_list = driver.find_element(By.XPATH, "//*[@id='code_ul']").find_elements(By.TAG_NAME, "li")
        code_id_value_list = []
        result_list = []
        try:
            for element in populated_list:
                code_id = element.find_element(By.TAG_NAME, "div").get_attribute('id')
                code_id = code_id.split("|")[0]
                code_value = element.find_element(By.TAG_NAME, "div").get_attribute('data-tooltip')
                code_desc = str(code_id) + " | " + str(code_value)
                code_id_value_list.append(code_id)
                code_id_value_list.append(code_value)
                result_list.append(code_desc)
            result_list = "\n ".join(result_list)
            driver.find_element(By.ID, "advanced_code_search_input").clear()
            driver.find_element(By.ID, "advanced_code_search_input").send_keys(code_id_value_list[0])
            time.sleep(5)
            populated_list = driver.find_element(By.XPATH, "//*[@id='code_ul']").find_elements(By.TAG_NAME, "li")
            result = []
            try:
                for element in populated_list:
                    code_id = element.find_element(By.TAG_NAME, "div").get_attribute('id')
                    code_id = code_id.split("|")[0]
                    code_value = element.find_element(By.TAG_NAME, "div").get_attribute('data-tooltip')
                    code_desc = str(code_id) + " | " + str(code_value)
                    result.append(code_desc)
                populated_result = "\n ".join(result)
            except NoSuchElementException as e:
                populated_result = "No Suggestions found for the searched code: " + str(code_id_value_list[0])
            status = get_status(code_id_value_list[0], result)
            try:
                result = "\n ".join(result)
                print(result)
                populated_result = code_id_value_list[0]+" | "+code_id_value_list[1]
            except TypeError as e:
                populated_result = "Check the result. Atleast one item empty."
        except NoSuchElementException as e:
            result_list = "No suggestions found"
            populated_result = "NA"
            status = "NA"
        ws.append([customer, lob, cz_id, measure_name, criteria, result_list, populated_result, status])
        driver.find_element(By.XPATH, "//*[@class='asf_header']//*[contains(text(),'clear')]").click()
        time.sleep(1)
        # driver.find_element(By.XPATH, "//*[@class='asf_header']//*[contains(text(),'clear')]").click()
        # driver.find_element(By.XPATH, "//*[@class='advanced_search_btn']").click()



def get_status(searched_value, returned_result_array):
    if not searched_value or not returned_result_array:
        return "Empty Result Set"

    s = str(searched_value).strip().lower()

    for item in returned_result_array:
        r = str(item).strip().lower()
        if s in r:
            return "MATCH"

    return "NO MATCH"


def format_excel_sheet(ws):
    header_fill = PatternFill("solid", fgColor="D9E1F2")
    header_font = Font(bold=True)
    center_align = Alignment(vertical="center", wrap_text=True)

    max_col = ws.max_column
    max_row = ws.max_row

    # Format header row
    for col in range(1, max_col + 1):
        cell = ws.cell(row=1, column=col)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = center_align

        column_letter = get_column_letter(col)
        ws.column_dimensions[column_letter].width = 30

    # Wrap text for Returned Result column
    for row in range(2, max_row + 1):
        ws.cell(row=row, column=4).alignment = Alignment(wrap_text=True)


    # Freeze header
    ws.freeze_panes = "A2"

    # Conditional formatting for Status column
    for row in range(2, max_row + 1):
        status_cell = ws.cell(row=row, column=8)

        if status_cell.value == "MATCH":
            status_cell.fill = PatternFill("solid", fgColor="C6EFCE")   # Green
        elif status_cell.value == "NO MATCH":
            status_cell.fill = PatternFill("solid", fgColor="F4CCCC")  # Red
        elif status_cell.value == "Empty Return Set":
            status_cell.fill = PatternFill("solid", fgColor="FFFFCC")  # Yellow

def url_navigator(cust_id, base_url, main_url):
    parsed_url = urlparse(main_url)
    query_parameters = parse_qs(parsed_url.query)
    session_id = query_parameters.get("session", [None])[0]
    decoded_val = base64.b64decode(session_id).decode('utf-8')
    temp = decoded_val.replace('1500', str(cust_id))
    session_id_temp = base64.b64encode(temp.encode('utf-8'))
    temp1 = str(session_id_temp.decode('utf-8'))
    link = base_url + "registries?session=" + temp1
    return link

'''INPUT PARAMS IN CODE'''

login_url = "https://www.cozeva.com/user/login"  #select environment
logout_url = "https://www.cozeva.com/user/logout"
base_url = "https://www.cozeva.com/"
User = os.environ.get('CS2_User')
Pass = os.environ.get('CS2_Password')
print(result)
if result == None:
    sys.exit(0)
else:
    measure_abb = result["abbrev"]
    measure_name = result["measure"]
    lob_cust_mapping = result["lob_customer_mapping"]



path1 = "C:\\Users\\SumanBasu\\Automation\\Report\\"
name_date = datetime.datetime.now()
formatted_date = name_date.strftime("%y-%m-%d %H-%M")
sheet_week = name_date.strftime("%m/%d")
filename = measure_abb +" validation_" + str(formatted_date) + ".xlsx"
wb = Workbook()
ws = wb.active
sheet_name = measure_abb + " Navigation"
ws.title = sheet_name
ws.append([
    "Customer Name",
    "LOB",
    "Domain Name check",
    "Measure Name check ",
    "Performance Statistics Score",
    "Network Comparison Score",
    "Comments"
])
format_excel_sheet(ws)
wb.create_sheet("MSPL vs Dashboard Check")
ws = wb["MSPL vs Dashboard Check"]
ws.append(["Customer Name","LOB","Patient Compliancy","Pencil icon present in MSPL", "MSPL pencil icon options", "Dashboard Pencil icon present?","MAP Checkbox Present?","MSPL vs Dashboard values", "MSPL link"])
format_excel_sheet(ws)
wb.create_sheet("Code Search Validation")
ws = wb["Code Search Validation"]
ws.append(["Customer", "LOB Name", "Cozeva ID","Measure Name","Criteria Name", "Populated codes", "Search Result Populated", "Result Matched"])
format_excel_sheet(ws)


#chrome-setup
options = webdriver.ChromeOptions()
prefs = {"download.default_directory" : "C:\\Users\\SumanBasu\\Automation\\CSV_Files"}
options.add_argument("user-data-dir=C:\\Users\\SumanBasu\\OneDrive - Vatica Health\\AppData\\Local\\Google\\Chrome\\User Data\\SavedData")
options.add_argument("--disable-notifications")
options.add_experimental_option("prefs", prefs)
driver = webdriver.Chrome(executable_path="C:\\Users\\SumanBasu\\Downloads\\RTVS_dist\\assets\\chromedriver.exe", options=options)


driver.get(logout_url)
driver.get(login_url)
driver.maximize_window()
print("Initial window= " + driver.title)
driver.find_element(By.ID, "edit-name").send_keys(User)
driver.find_element(By.ID, "edit-pass").send_keys(Pass)
driver.find_element(By.ID, "edit-submit").click()
time.sleep(1)
Last_access_check_flag = 0
try:
    WebDriverWait(driver, 90).until(EC.presence_of_element_located((By.ID, "reason_textbox")))
    driver.find_element(By.ID, "reason_textbox").send_keys("https://redmine2.cozeva.com/issues/48191")
    driver.find_element(By.ID, "edit-submit").click()
except NoSuchElementException:
    traceback.print_exc()
    driver.quit()
print("Landing page= " + driver.title)
ajax_preloader_wait(driver)
if driver.title != "Registries | Cozeva":
    driver.find_element(By.XPATH, "//*[@data-target='app_dropdown']").click()
    driver.find_element(By.XPATH, "//*[@class='no-hover app_registries']").click()
    driver.switch_to.window(driver.window_handles[1])
ajax_preloader_wait(driver)
registry_url = driver.current_url
print("Starting validaton of measure: "+measure_abb)
measure_domain = "NA"
for record in lob_cust_mapping:
    lob = record["lob"]
    customer_name = record["customer"]
    customer_id = record["customer_id"]
    measure_domain = record["domain"]
    measure_year = "2026"
    print("Checking "+measure_abb+" for "+customer_name+" in "+lob+" of MY2026")
    link = url_navigator(customer_id, base_url, registry_url)
    driver.get(link)
    ajax_preloader_wait(driver)
    driver.find_element(By.XPATH, "//*[@id='qt-filter-label']").click()
    time.sleep(1)
    years = driver.find_element(By.XPATH, "//*[@id='filter-quarter']").find_elements(By.TAG_NAME, "li")
    for year in years:
        if measure_year in year.text:
            year.click()
    LOB_list = driver.find_element(By.XPATH, "//*[@id='filter-lob']").find_elements(By.TAG_NAME, 'li')
    for LOB in LOB_list:
        if "All" in lob:
            if lob in LOB.text:
                LOB.click()
                break
        else:
            if lob == LOB.text:
                LOB.click()
                break
    time.sleep(1)
    driver.find_element(By.XPATH, "//*[@id='reg-filter-apply']").click()
    ajax_preloader_wait(driver)
    ws = wb[sheet_name]
    try:
        measure_search_registry(driver, measure_name)
        code_check_flag = measure_patient_finder(driver, customer_name, lob, measure_name, measure_abb, measure_domain, ws, wb)
        wb.save(path1+filename)
        if code_check_flag == 1:
            print("Supplemental data can be added, proceeding to code search...")
            supplemental_data_launcher(driver)
            ws = wb["Code Search Validation"]
            code_search_coding_tool(driver, customer_name, lob, measure_name, ws)
            format_excel_sheet(ws)
            wb.save(path1+filename)
            driver.find_element(By.XPATH, "//*[@id='chart_action_82']").click()
            WebDriverWait(driver, 30).until(EC.visibility_of_element_located((By.XPATH, "//*[contains(@class,'cozeva-prompt')]")))
            time.sleep(1)
            driver.find_element(By.XPATH, "//*[@id='task_delete_input']").send_keys("CozevaQA")
            driver.find_element(By.XPATH, "//*[@data-index='confirm']").click()
            time.sleep(1)
            ajax_preloader_wait(driver)
        else:
            print("No pencil icon available, skipping to next LOB")
    except Exception as e:
        print(e)

driver.close()


# Multiple code search button
# Phase 2: MSPL column blank check, Relevant care history check, Submission of data in Simulated Customer
